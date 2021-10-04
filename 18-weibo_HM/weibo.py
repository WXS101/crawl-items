import json
import xlrd
import requests
import re
import time
from lxml import etree
from openpyxl import Workbook
import pymongo
from matplotlib import pyplot as plt
from wordcloud import WordCloud
from PIL import Image
import numpy as np
import jieba
from collections import Counter


def get_data(url):
    head = {
        "MWeibo-Pwa": "1",
        "Referer": "https://m.weibo.cn/search?containerid=100103type%3D1%26q%3DHM",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36"
    }
    response = requests.get(url, headers=head)
    response = response.json()
    print(response)
    # data 的长度为10,相当于一页10个内容i
    data = response['data']['cards']
    for j in range(len(data)):
        try:
            result = data[j]['mblog']['text']
            res = etree.HTML(result)
            a = "".join(res.xpath('//text()'))
            data_list.append([a])
        #     collection.insert_one({
        #         "time": otherStyleTime
        #     })
        except KeyError as p:
            print("错误")
            continue
    print(data_list)


def wordcloud(title):
    # 分词
    cut = jieba.cut(title)
    string = " ".join(cut)
    # 打开遮罩图片
    img = Image.open(r"prcie.jpg")
    # 将图片转换为数组
    img_array = np.array(img)
    # 设置参数
    wc = WordCloud(
        background_color='white',
        mask=img_array,
        font_path="msyh.ttc"
    )
    wc.generate_from_text(string)
    fig = plt.figure(1)
    plt.imshow(wc)
    # 是否显示坐标轴
    plt.axis('off')
    # 显示生成的词云图片,可以直接保存就注释了
    # plt.show()
    # 保存
    plt.savefig(r'wordcloud.jpg', dpi=500)


def statistics_word(title):
    cut = jieba.cut(title)
    cutList = []
    # 分词
    for word in cut:
        if (word in title) and len(word) > 1:
            cutList.append(word)
    wordList = []
    # 统计每个词出现的次数
    wordList = Counter(cutList)
    # 统计出现的次数最多的N个词及出现的次数
    wordListNTop = wordList.most_common(10)
    # 保存到excel里面
    file = Workbook()
    sheet = file.active
    # 设置标题
    sheet.append(["排名", "出现的次数"])
    for i in wordListNTop:
        sheet.append(i)
    file.save('cipin.xlsx')


if __name__ == '__main__':
    # mongo_py = pymongo.MongoClient()
    #
    # collection = mongo_py['weibo_HM']['data']
    data_list = []
    for i in range(1):
        try:
            time.sleep(1)
            url = "https://m.weibo.cn/api/container/getIndex?containerid=2304131866405545_-_WEIBO_SECOND_PROFILE_WEIBO&luicode=10000011&lfid=2302831866405545&page_type=03&page="+str(i)+""
            print(url)
            get_data(url)
        except:
            continue

    # 词云
    # wb = xlrd.open_workbook("微博.xlsx")
    # sh = wb.sheet_by_name('Sheet')
    # str_res = ""
    # for number in range(sh.nrows):
    #     str_res += sh.cell(number+1, 0).value
    # print(str_res)
    # # 调用保存词云函数
    # wordcloud(str_res)
    # # 调用统计词频函数
    # statistics_word(str_res)

    # fire = Workbook()
    # sheet = fire.active
    # sheet.append(["content"])
    # for m in data_list:
    #     sheet.append(m)
    # fire.save('微博_时间.xlsx')

    mongo_py = pymongo.MongoClient()

    collection = mongo_py['weibo_HM']['data']

    time_list = []

    for item in collection.find():
        if "04-03" in item['time']:
            time_list.append(item['time'].split()[1].split(":")[0])

    unique_data = np.unique(time_list)
    print(unique_data)
    resdata = []
    for ii in unique_data:
        resdata.append(time_list.count(ii))
    print(resdata)
    plt.figure(figsize=(12, 8))
    rects = plt.bar(unique_data,resdata,width=0.8)
    plt.savefig("直方图2.jpg")
