import json
import xlrd
import requests
import time
from lxml import etree
from openpyxl import Workbook
from matplotlib import pyplot as plt
from wordcloud import WordCloud
from PIL import Image
import numpy as np
import jieba
from collections import Counter


def get_data(url):
    # 定义请求头
    head = {
        "MWeibo-Pwa": "1",
        "Referer": "https://m.weibo.cn/search?containerid=100103type%3D1%26q%3DHM",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36"
    }
    # 获取响应
    response = requests.get(url, headers=head)
    # 转换为json
    response = response.json()
    print(response)
    # 拿到cards里面的数据
    data = response['data']['cards']
    # 遍历
    for j in range(len(data)):
        # 防止报错
        try:
            # 拿到text里面的数据
            result = data[j]['mblog']['text']
            # 使用xpath获取具体的数据
            res = etree.HTML(result)
            # 处理数据
            a = "".join(res.xpath('//text()'))
            # 添加到datalist列表里面
            data_list.append([a])
        except KeyError as p:
            print("错误")
            continue
    # 打印一下,查看进度,也可以不打印
    print(data_list)


def wordcloud(title):
    # 分词
    cut = jieba.cut(title)
    string = " ".join(cut)
    # 打开遮罩图片
    img = Image.open(r"prcie.jpg")
    # 将图片转换为数组
    img_array = np.array(img)
    # 设置参数
    wc = WordCloud(
        background_color='white',
        mask=img_array,
        font_path="msyh.ttc"
    )
    wc.generate_from_text(string)
    fig = plt.figure(1)
    plt.imshow(wc)
    # 是否显示坐标轴
    plt.axis('off')
    # 显示生成的词云图片,可以直接保存就注释了
    # plt.show()
    # 保存
    plt.savefig(r'wordcloud.jpg', dpi=500)


def statistics_word(title):
    cut = jieba.cut(title)
    cutList = []
    # 分词
    for word in cut:
        if (word in title) and len(word) > 1:
            cutList.append(word)
    wordList = []
    # 统计每个词出现的次数
    wordList = Counter(cutList)
    # 统计出现的次数最多的N个词及出现的次数
    wordListNTop = wordList.most_common(10)
    # 保存到excel里面
    file = Workbook()
    sheet = file.active
    # 设置标题
    sheet.append(["排名", "出现的次数"])
    # 遍历
    for i in wordListNTop:
        sheet.append(i)
    file.save('cipin.xlsx')



if __name__ == '__main__':
    # 定义一个列表,存储数据
    data_list = []
    # 循环爬取,这里是循环110页
    for i in range(155):
        try:
            # 停一秒,防止被封ip
            time.sleep(1)
            # 爬取的网址
            url = "https://m.weibo.cn/api/container/getIndex?containerid=2304131866405545_-_WEIBO_SECOND_PROFILE_WEIBO&luicode=10000011&lfid=2302831866405545&page_type=03&page="+str(i)+""
            # 打印一下网址,查看进度,也可以不打印
            print(url)
            # 调用函数
            get_data(url)
        except:
            continue

    # 使用Workbook保存数据到 微博_谣言.xlsx文件中
    fire = Workbook()
    sheet = fire.active
    sheet.append(["content"])
    for m in data_list:
        sheet.append(m)
    fire.save('微博_谣言2.xlsx')

    # 词云
    wb = xlrd.open_workbook("微博_谣言.xlsx")
    # 定位到Sheet
    sh = wb.sheet_by_name('Sheet')
    str_res = ""
    for number in range(sh.nrows):
        # 处理数据,方便做词云
        str_res += sh.cell(number, 0).value
    print(str_res)
    # 调用保存词云函数
    wordcloud(str_res)
    # 调用统计词频函数
    statistics_word(str_res)
