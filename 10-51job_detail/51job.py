import requests
import xlrd
import re
from lxml import etree
import json
import time
import pymongo
from openpyxl import Workbook

data_list = []
url_list = []

mongo_py = pymongo.MongoClient()

collection = mongo_py['51job_kuaiji']['data']


def get_response(url):
    head = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36"
    }
    response = requests.get(url, headers=head)
    if response.status_code == 200:
        return response.text
    else:
        print("访问网页错误")


# 爬取详情页的数据
def parse_data(url):
    try:
        job_data = []
        head = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36"
        }
        response = requests.get(url, headers=head).content.decode('gbk')
        data = etree.HTML(response)
        job_name = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/h1/text()')
        job_name = "没有" if job_name == [] else job_name[0]
        price = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/strong/text()')
        price = "没有" if price == [] else price[0]
        company_name = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[1]/a[1]/text()')[0]
        if len(data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')) == 5 or len(
                data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')) > 5:
            issue_time = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[4].replace("  ", "")
            address = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[0].replace("  ", "")
            experience = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[1].replace("  ", "")
            education = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[2].replace("  ", "")
            people_number = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[3].replace("  ", "")
        else:
            address = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[0].replace("  ", "")
            issue_time = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[-1].replace("  ", "")
            for j in range(len(data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()'))):
                if "人" in data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[j]:
                    people_number = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[j].replace(
                        "  ",
                        "")
                else:
                    people_number = "没有"
            for m in range(len(data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()'))):
                if "经验" in data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[m]:
                    experience = data.xpath('/html/body/div[3]/div[2]/div[2]/div/div[1]/p[2]/text()')[m].replace("  ",
                                                                                                                 "")
                else:
                    experience = "没有"
            education = "没有"
        company_type = data.xpath('/html/body/div[3]/div[2]/div[4]/div[1]/div[2]/p[1]/text()')
        company_type = "没有" if company_type == [] else company_type[0]
        company_people = data.xpath('/html/body/div[3]/div[2]/div[4]/div[1]/div[2]/p[2]/text()')
        company_people = "没有" if company_people == [] else company_people[0]
        company_item = data.xpath('/html/body/div[3]/div[2]/div[4]/div[1]/div[2]/p[3]/a/text()')
        company_item = "没有" if company_item == [] else company_item[0]
        position_information = data.xpath('/html/body/div[3]/div[2]/div[3]/div[1]/div//text()')
        position_information = \
        "".join(position_information).strip().split("\n", 0)[0].split("职能类别", 1)[0].replace("\r\n", "").split("工作时间",
                                                                                                              1)[
            0].split("联系人", 1)[0].replace("\xa0", "").replace("\u3000\u3000", "")
        # data_list.append(
        #     [job_name, price, company_name, issue_time, address, experience, education, people_number, company_type,
        #      company_people, company_item, position_information])
        collection.insert_one({
            "工作名称": job_name,
            "薪资": price,
            "公司名称": company_name,
            "发布时间": issue_time,
            "地址": address,
            "经验要求": experience,
            "学历要求": education,
            "招聘人数": people_number,
            "公司类型": company_type,
            "公司人数": company_people,
            "公司项目": company_item,
            "职位信息": position_information
        })
    except IndexError as e:
        print(e)


def main():
    try:
        for i in range(301, 491):
            print(i)
            out_url = 'https://search.51job.com/list/000000,000000,0000,00,9,99,%25E8%25B4%25A2%25E5%258A%25A1%25E4%25BC%259A%25E8%25AE%25A1,2,' + str(
                i) + '.html?lang=c&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare='
            response = get_response(out_url)
            # 用正则爬取
            data = re.findall(r'"engine_search_result":(.*?),"jobid_count"', response)[0]
            data = json.loads(data)
            for item in data:
                url_list.append([item["job_href"]])
    except UnicodeDecodeError as e:
        print("错误")

    fire = Workbook()
    sheet = fire.active
    for m in url_list:
        sheet.append(m)
    fire.save('51job_url.xlsx')

    wb = xlrd.open_workbook("51job_url.xlsx")
    sh = wb.sheet_by_name('Sheet')
    for i in range(8956, 9482):
        print(i)
        print(sh.cell(i, 0).value)
        a = sh.cell(i, 0).value
        if "https" in a:
            parse_data(a)
        else:
            continue

    datas_list = []
    for item in collection.find():
        datas_list.append(
            [item['工作名称'], item['薪资'], item['公司名称'], item['发布时间'], item['地址'], item['经验要求'], item['学历要求'], item['招聘人数'],
             item['公司类型'], item['公司人数'], item['公司项目'], item['职位信息']])
    fire = Workbook()
    sheet = fire.active
    sheet.append(['工作名称', '薪资', '公司名称', '发布时间', '地址', '经验要求', '学历要求', '招聘人数', '公司类型', '公司人数', '公司项目', '职位信息'])
    for i in datas_list:
        sheet.append(i)
    fire.save('51job_url.xlsx')


if __name__ == '__main__':
    main()
